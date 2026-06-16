"""Testy parsera KPiR na próbce SMAKOSZ (kwiecień 2026)."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from audytor.parser_kpir import ParserError, _czytaj_wiersz_sum, parse_kpir
from audytor.patterns import okres_dra, okres_listy_plac, proporcja_paliwa

FIXTURE = Path(__file__).parent / "fixtures" / "kpir_smakosz_2026_04.xlsx"

wymaga_fixture = pytest.mark.skipif(
    not FIXTURE.exists(),
    reason="Brak realnego pliku KPiR (dane klienta, poza repo) — umieść go w tests/fixtures/",
)


@pytest.fixture(scope="module")
def ksiega():
    return parse_kpir(FIXTURE)


class TestOdpornoscNaPusteWiersze:
    def test_pusty_wiersz_nie_wywala_index_error(self):
        # openpyxl (read_only) na niektórych wersjach zwraca puste wiersze
        # jako puste krotki — dostęp do row[0] musi być bezpieczny.
        rows = [(), ("Suma miesiąca:", "172476.44"), ()]
        sumy = _czytaj_wiersz_sum(rows, "Suma miesiąca:", {11: 1})
        assert sumy[11] == Decimal("172476.44")


@wymaga_fixture
class TestParsowanieKsiegi:
    def test_naglowek(self, ksiega):
        assert ksiega.nip_podatnika == "7162519569"
        assert (ksiega.rok, ksiega.miesiac) == (2026, 4)
        assert "SMAKOSZ" in ksiega.nazwa_podatnika

    def test_liczba_wpisow(self, ksiega):
        assert len(ksiega.wpisy) == 136

    def test_sumy_zgodne_z_podsumowaniem(self, ksiega):
        # razem przychód (11) i razem wydatki (16) — kotwice z wydruku
        assert ksiega.suma_miesiaca[11] == Decimal("172476.44")
        assert ksiega.suma_miesiaca[16] == Decimal("82240.19")
        assert ksiega.narastajaco[11] == Decimal("708106.42")

    def test_pierwszy_wpis_kompletny(self, ksiega):
        wpis = ksiega.wpisy[0]
        assert wpis.lp == 1
        assert wpis.data == date(2026, 4, 1)
        assert wpis.identyfikator == "7160005777"
        assert wpis.opis == "zakup towaru"
        assert wpis.kontrahent.startswith("WIESŁAW TRYBIŁO")
        assert wpis.nr_dowodu == "0171/02/2026"
        assert wpis.kwota(12) == Decimal("479.26")

    def test_nr_ksef(self, ksiega):
        wpis = next(w for w in ksiega.wpisy if w.lp == 3)
        assert wpis.nr_ksef == "7791906082-20260401-3BE8CB0000FD-43"


@wymaga_fixture
class TestRozpoznawanieWpisow:
    def test_proporcje_paliwa(self, ksiega):
        proporcje = {proporcja_paliwa(w.opis) for w in ksiega.wpisy if proporcja_paliwa(w.opis)}
        assert proporcje == {20, 75}

    def test_lista_plac(self, ksiega):
        lpu = next(w for w in ksiega.wpisy if w.lp == 39)
        lpe = next(w for w in ksiega.wpisy if w.lp == 40)
        assert okres_listy_plac(lpu.nr_dowodu) == ("LPU", 2026, 3)
        assert okres_listy_plac(lpe.nr_dowodu) == ("LPE", 2026, 3)
        assert lpu.kwota(14) == Decimal("32679.00")

    def test_dra(self, ksiega):
        dra = next(w for w in ksiega.wpisy if w.lp == 77)
        assert okres_dra(dra.nr_dowodu) == (2026, 2)

    def test_incydentalny(self, ksiega):
        incydentalne = [w for w in ksiega.wpisy if w.incydentalny]
        assert len(incydentalne) == 2

    def test_korekta_ujemna(self, ksiega):
        korekta = next(w for w in ksiega.wpisy if w.lp == 101)
        assert korekta.kwota(12) == Decimal("-42.84")


@wymaga_fixture
class TestBledyParsowania:
    def test_zepsuta_kwota_rzuca_parser_error(self, tmp_path):
        workbook = openpyxl.load_workbook(FIXTURE)
        arkusz = workbook.active
        # wiersz 18 = pierwszy wpis, kolumna 13 = (12) zakup towarów: 479,26
        assert arkusz.cell(row=18, column=13).value == "479,26"
        arkusz.cell(row=18, column=13).value = "999,99"
        zepsuty = tmp_path / "zepsuty.xlsx"
        workbook.save(zepsuty)

        with pytest.raises(ParserError, match="Niezgodność sum"):
            parse_kpir(zepsuty)

    def test_plik_bez_naglowka_kpir(self, tmp_path):
        workbook = openpyxl.Workbook()
        workbook.active.cell(row=1, column=1).value = "to nie jest KPiR"
        inny = tmp_path / "inny.xlsx"
        workbook.save(inny)

        with pytest.raises(ParserError):
            parse_kpir(inny)
